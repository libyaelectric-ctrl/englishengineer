import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Yeni C2 terimlerini yükle
new_all = pd.read_json(r'C:\Users\User\Desktop\EngineerOS_DENEME_CODEX\8.0\analysis_output\genel8.json')

# Excel dosyasını oluştur
wb = openpyxl.Workbook()

# Özet sayfası (Summary)
ws_summary = wb.active
ws_summary.title = "Ozet"

# Başlık stili
header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True, size=12)
title_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
title_font = Font(bold=True, size=14)

ws_summary['A1'] = "GENEL8 - CEFR GAP DOLDURMA SONUCU"
ws_summary['A1'].font = title_font
ws_summary['A1'].fill = title_fill

ws_summary['A3'] = "Orijinal Veri"
ws_summary['B3'] = "Genel7.xlsx (14,622 kayıt)"
ws_summary['A4'] = "Temizlenmiş Veri"
ws_summary['B4'] = "genel7_final.csv (14,646 kayıt)"
ws_summary['A5'] = "Yeni C2 Terim (gene8)"
ws_summary['B5'] = "genel8.json (24 kayıt)"

ws_summary['A7'] = "Industrial C2"
ws_summary['B7'] = "Önce: 6 | Sonra: 20 (+14 yeni)"
ws_summary['A8'] = "Mechatronics C2"
ws_summary['B8'] = "Önce: 5 | Sonra: 15 (+10 yeni)"

ws_summary['A10'] = "Toplam Yeni Terim"
ws_summary['B10'] = 24

# Kolon genişlikleri
ws_summary.column_dimensions['A'].width = 25
ws_summary.column_dimensions['B'].width = 35

# Yeni terimler sayfası
ws_data = wb.create_sheet("C2_Gap_Terimleri")

# Başlık satırı
headers = ["id", "term", "normalizedTerm", "turkishMeaning", "cefrLevel", "domain",
           "contentDomain", "category", "termType", "partOfSpeech", "wordCount",
           "definition", "exampleSentence", "turkishExample", "tags", "confidence", "status"]

for col, header in enumerate(headers, 1):
    cell = ws_data.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

# Verileri yaz
for row_idx, (_, row) in enumerate(new_all.iterrows(), 2):
    for col_idx, header in enumerate(headers, 1):
        value = row.get(header, "")
        cell = ws_data.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")
        # CEFR seviyesine göre renk
        if header == "cefrLevel" and str(value) == "C2":
            cell.font = Font(bold=True, color="FF0000")
            cell.fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
        # Domain renklendirme
        if header == "domain" and str(value) == "industrial":
            cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        elif header == "domain" and str(value) == "mechatronics":
            cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        cell.alignment = Alignment(wrap_text=True, vertical='top')

# Kolon genişlikleri
ws_data.column_dimensions['A'].width = 30  # id
ws_data.column_dimensions['B'].width = 35  # term
ws_data.column_dimensions['C'].width = 35  # normalizedTerm
ws_data.column_dimensions['D'].width = 30  # turkishMeaning
ws_data.column_dimensions['E'].width = 10  # cefrLevel
ws_data.column_dimensions['F'].width = 15  # domain
ws_data.column_dimensions['G'].width = 15  # contentDomain
ws_data.column_dimensions['H'].width = 25  # category
ws_data.column_dimensions['I'].width = 15  # termType
ws_data.column_dimensions['J'].width = 15  # partOfSpeech
ws_data.column_dimensions['K'].width = 10  # wordCount
ws_data.column_dimensions['L'].width = 60  # definition
ws_data.column_dimensions['M'].width = 50  # exampleSentence
ws_data.column_dimensions['N'].width = 35  # turkishExample
ws_data.column_dimensions['O'].width = 30  # tags
ws_data.column_dimensions['P'].width = 8   # confidence
ws_data.column_dimensions['Q'].width = 10  # status

# Satır yükseklikleri
for row in range(2, len(new_all)+2):
    ws_data.row_dimensions[row].height = 60

# Otomatik filtre ekle
ws_data.auto_filter.ref = f"A1:Q{len(new_all)+1}"

# Üçüncü sayfa: Tüm sonuç dosyaları listesi (rapor)
ws_report = wb.create_sheet("Rapor")
ws_report['A1'] = "GENEL8 - ANALİZ SONUÇ RAPORU"
ws_report['A1'].font = title_font
ws_report['A1'].fill = title_fill

ws_report['A3'] = "İşlem"
ws_report['B3'] = "Durum"
ws_report['C3'] = "Detay"

for col in [1,2,3]:
    cell = ws_report.cell(row=3, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

rapor_veriler = [
    ("Genel7.xlsx Okuma", "✓ Tamamlandı", "14,622 kayıt okundu"),
    ("Duplicate Temizleme", "✓ Tamamlandı", "0 kayıt çıkarıldı (benzersizdi)"),
    ("Cross-Domain Duplicate", "✓ Tespit Edildi", "414 terim farklı domainlerde tekrar ediyor"),
    ("CEFR Gap - Industrial C2", "✓ Dolduruldu", "+14 yeni C2 terimi eklendi"),
    ("CEFR Gap - Mechatronics C2", "✓ Dolduruldu", "+10 yeni C2 terimi eklendi"),
    ("genel8.json Oluşturma", "✓ Tamamlandı", "24 yeni C2 terimi"),
    ("gene8_summary.txt", "✓ Tamamlandı", "Özet rapor yazıldı"),
]

for i, (islem, durum, detay) in enumerate(rapor_veriler, 5):
    ws_report['A'+str(i)] = islem
    ws_report['B'+str(i)] = durum
    ws_report['C'+str(i)] = detay
    ws_report['B'+str(i)].font = Font(bold=True, color="008000")

ws_report.column_dimensions['A'].width = 35
ws_report.column_dimensions['B'].width = 18
ws_report.column_dimensions['C'].width = 35

# Excel'i kaydet
output_path = r'C:\Users\User\Desktop\EngineerOS_DENEME_CODEX\database\dosyalar\genel8.xlsx'
wb.save(output_path)
print(f"genel8.xlsx oluşturuldu: {output_path}")
print(f"Yeni C2 terim sayısı: {len(new_all)}")
